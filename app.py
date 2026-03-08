"""
NjoPerKrejt - SmartRegister
Version manuale - pa AI, pa internet
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG FILE (ruhet prane app.py) ──────────────────────────────────────────
APP_DIR     = Path(os.path.expanduser("~")) / "NjoPerKrejt"
APP_DIR.mkdir(parents=True, exist_ok=True)
CONFIG_FILE = APP_DIR / "config.json"

VERSION = "1.0"
GITHUB_USER = "riadspahiu"
GITHUB_REPO = "njoperkrejt-app"
GITHUB_BRANCH = "main"
GITHUB_RAW = f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/{GITHUB_BRANCH}"


def check_for_update():
    """Kontrollo GitHub per version te ri. Ekzekutohet ne background thread."""
    import threading

    def _check():
        try:
            import urllib.request, urllib.error
            url = f"{GITHUB_RAW}/version.txt"
            with urllib.request.urlopen(url, timeout=5) as r:
                latest = r.read().decode().strip()
            if latest and latest != VERSION:
                _prompt_update(latest)
        except Exception:
            pass  # nuk ka internet ose repo — vazhdo normalisht

    threading.Thread(target=_check, daemon=True).start()


def _prompt_update(latest_version):
    """Shfaq dritaren e update-it ne thread kryesor."""
    import urllib.request

    def _show():
        win = tk.Toplevel()
        win.title("Update i disponueshem")
        win.geometry("380x180")
        win.configure(bg="#0f1117")
        win.resizable(False, False)
        win.grab_set()
        win.lift()
        win.focus_force()

        tk.Label(win, text="🔄  Update i ri i disponueshem!",
                 font=("Segoe UI", 11, "bold"), bg="#0f1117", fg="#66FFCC").pack(pady=(22, 4))
        tk.Label(win, text=f"Versioni aktual: {VERSION}   →   Versioni i ri: {latest_version}",
                 font=("Segoe UI", 9), bg="#0f1117", fg="#888").pack()
        tk.Label(win, text="Deshiron ta instalosh tani?",
                 font=("Segoe UI", 9), bg="#0f1117", fg="#ccc").pack(pady=(8, 16))

        btn_row = tk.Frame(win, bg="#0f1117")
        btn_row.pack()

        def do_update():
            try:
                url = f"{GITHUB_RAW}/app.py"
                dest = Path(sys.executable).parent / "app.py"
                # Nese eshte frozen (PyInstaller), shiko ne _MEIPASS ose pranë exe
                if getattr(sys, 'frozen', False):
                    dest = Path(sys.executable).parent / "app.py"
                urllib.request.urlretrieve(url, dest)
                win.destroy()
                # Rihap programin
                import subprocess
                subprocess.Popen([sys.executable] + sys.argv)
                import os; os._exit(0)
            except Exception as e:
                tk.messagebox.showerror("Gabim", f"Update deshtoi:\n{e}", parent=win)

        tk.Button(btn_row, text="  Po, instalo  ", font=("Segoe UI", 9, "bold"),
                  bg="#66FFCC", fg="#0f1117", relief="flat", padx=16, pady=8,
                  cursor="hand2", command=do_update).pack(side="left", padx=(0, 10))
        tk.Button(btn_row, text="Jo tani", font=("Segoe UI", 9),
                  bg="#1a1d26", fg="#888", relief="flat", padx=16, pady=8,
                  cursor="hand2", command=win.destroy).pack(side="left")

    # Thirre ne thread kryesor pas 1 sekonde (sigurohu qe App() eshte gati)
    try:
        tk._default_root.after(1000, _show)
    except Exception:
        pass

def load_config():
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except:
            pass
    return {}

def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")

def get_base_dir():
    cfg = load_config()
    p = cfg.get("base_dir", "")
    if p and Path(p).exists():
        return Path(p)
    return APP_DIR

def get_biz_dir():
    """Kthen folder-in e bizneseve — mund te jete direkt folder-i i klienteve."""
    cfg = load_config()
    p = cfg.get("biz_dir", "")
    if p and Path(p).exists():
        return Path(p)
    return get_base_dir() / "Bizneset"

def set_base_dir(path):
    cfg = load_config()
    cfg["base_dir"] = str(path)
    save_config(cfg)

def set_dirs(base_path, biz_path=None):
    """Ruan base_dir dhe opsionalisht biz_dir te ndare."""
    cfg = load_config()
    cfg["base_dir"] = str(base_path)
    if biz_path and str(biz_path) != str(base_path):
        cfg["biz_dir"] = str(biz_path)
    else:
        cfg.pop("biz_dir", None)
    save_config(cfg)

SYS_FOLDER = "NjoPerKrejt - Excel [SpahiuDev]"

# ── DYNAMIC PATHS (rikomputon cdo here) ───────────────────────────────────────
def paths():
    biz  = get_biz_dir()
    base = biz / SYS_FOLDER
    lg   = base / "log.txt"
    for d in [base, biz]:
        d.mkdir(parents=True, exist_ok=True)
    return base, biz, lg

# ── COLORS — Night Mode Neutral ───────────────────────────────────────────────
BG      = "#0d0f14"
SURFACE = "#13151c"
CARD    = "#181b23"
BORDER  = "#252830"
ACCENT  = "#66FFCC"
GREEN   = "#22c55e"
YELLOW  = "#eab308"
RED     = "#ef4444"
TEXT    = "#e8eaf0"
MUTED   = "#4b5468"
WHITE   = "#f1f5f9"
H_BG    = "1a3a5c"
ROW_ALT = "f0f4f8"
INPUT_BG    = "#1e2130"
INPUT_FG    = "#f1f5f9"
INPUT_FOCUS = "#66FFCC"

FONT   = ("Segoe UI", 13)
FONT_B = ("Segoe UI", 13, "bold")
FONT_S = ("Segoe UI", 11)
FONT_H = ("Segoe UI", 14, "bold")

# ── LOGGING ───────────────────────────────────────────────────────────────────
def log(msg):
    _, _, lg = paths()
    ts = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    with open(lg, "a", encoding="utf-8") as f:
        f.write(f"[{ts}]  {msg}\n")

# ── EXCEL ─────────────────────────────────────────────────────────────────────
HEADERS_BIZNES  = ["Data", "Ora", "Produkti/Sherbimi", "Sasia", "Njesia",
                   "Cmimi/Njesi (EUR)", "Pagesa (EUR)", "Totali (EUR)", "Shenime"]
def style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=H_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right=Side(style="thin",   color="CCCCCC"))
    ws.row_dimensions[1].height = 28

def auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 35)

def ensure_biznes(name):
    _, biz, _ = paths()
    path = biz / f"{name.strip().replace(' ', '_')}.xlsx"
    if not path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dergesat"
        style_header(ws, HEADERS_BIZNES)
        wb.save(path)
        log(f"File i ri u krijua: {path.name}")
    return path

def append_row(path, row_data, headers):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    if ws.max_row == 0 or ws.cell(1, 1).value != headers[0]:
        style_header(ws, headers)
    nr = ws.max_row + 1
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=nr, column=col, value=val)
        cell.alignment = Alignment(vertical="center")
        if nr % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=ROW_ALT)
        cell.border = Border(bottom=Side(style="hair", color="CCCCCC"))
    auto_width(ws)
    wb.save(path)

# ── COLUMN DETECTION & SMART MAPPING ─────────────────────────────────────────
FIELD_ALIASES = {
    "date":     ["data", "date", "dt", "datum", "dita"],
    "ora":      ["ora", "koha", "time", "ore", "hour"],
    "biznesi":  ["biznesi", "bizneset", "klienti", "kliente", "firma",
                 "kompania", "emri", "business", "client", "name"],
    "produkti": ["produkti", "produkte", "sherbimi", "sherbime", "artikulli",
                 "artikuj", "product", "service", "item", "pershkrimi",
                 "description", "mall"],
    "sasia":    ["sasia", "quantity", "qty", "sasi"],
    "njesia":   ["njesia", "njesi", "unit", "measure"],
    "cmimi":    ["cmimi", "price", "cmim", "cena", "tarifa",
                 "cmimi/njesi", "price/unit", "cmimi per njesi"],
    "pagesa":   ["pagesa", "payment", "paid", "paguar",
                 "pagesa e marre", "total paguar"],
    "total":    ["total", "totali", "shuma", "gjithsej",
                 "totali i fatures", "grand total"],
    "shenime":  ["shenime", "notes", "note", "koment", "remarks", "info"],
}

def detect_column_mapping(xlsx_path):
    """Lexon headers e file-it dhe kthen mapping { field: col_index_1based }"""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        first_row = [str(c.value or "").strip().lower()
                     for c in next(ws.iter_rows(min_row=1, max_row=1))]
        original  = [str(c.value or "").strip()
                     for c in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
    except:
        return {k: None for k in FIELD_ALIASES}, []

    mapping = {}
    for field, aliases in FIELD_ALIASES.items():
        found = None
        for idx, hdr in enumerate(first_row):
            if hdr in aliases:
                found = idx + 1
                break
        mapping[field] = found
    return mapping, original

def scan_folder_columns(folder_path):
    """Skanon te gjitha .xlsx ne folder dhe kthen mappings"""
    result = {}
    for f in Path(folder_path).glob("*.xlsx"):
        if SYS_FOLDER in str(f):
            continue  # Kapërcen file-at e sistemit
        mapping, headers = detect_column_mapping(f)
        result[f.name] = {"mapping": mapping, "headers": headers, "path": f}
    return result

def append_row_mapped(path, entry_dict, mapping):
    """Shton rresht duke respektuar kolonat ekzistuese te file-it.
    Nuk shton kolona te reja — vetem ploteson ato qe gjenden.
    Nese add_timestamp=True ne config, shton kolonën Data/Ora ne fund (vetem here e pare).
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    nr = ws.max_row + 1
    max_col = ws.max_column or 0

    for field, col_idx in mapping.items():
        if col_idx is None:
            continue
        val = entry_dict.get(field, "")
        if val is None:
            val = ""
        cell = ws.cell(row=nr, column=col_idx, value=val)
        cell.alignment = Alignment(vertical="center")
        if nr % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=ROW_ALT)
        cell.border = Border(bottom=Side(style="hair", color="CCCCCC"))

    # Shto Data/Ora vetem nese perdoruesi ka zgjedhur "Po"
    cfg = load_config()
    if cfg.get("add_timestamp", False):
        ts_label = "Data/Ora e shtimit"
        # Kontrolloj nese kolona ekziston tashmë
        ts_col = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=c).value or "").strip() == ts_label:
                ts_col = c
                break
        if ts_col is None:
            # Krijo kolonën herën e parë
            ts_col = ws.max_column + 1
            hcell = ws.cell(row=1, column=ts_col, value=ts_label)
            hcell.font = Font(bold=True, color="FFFFFF", size=11)
            hcell.fill = PatternFill("solid", fgColor="1e3a5f")
            hcell.alignment = Alignment(horizontal="center", vertical="center")
        ts_val = entry_dict.get("date", "") + "  " + entry_dict.get("ora", "")
        cell = ws.cell(row=nr, column=ts_col, value=ts_val.strip())
        cell.alignment = Alignment(vertical="center")
        if nr % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=ROW_ALT)

    auto_width(ws)
    wb.save(path)

def save_entry(entry):
    b   = entry["biznesi"].strip()
    now = datetime.now()
    dt  = entry["date"] or now.strftime("%d/%m/%Y")
    ora = now.strftime("%H:%M:%S")

    full_entry = dict(entry)
    full_entry["date"] = dt
    full_entry["ora"]  = ora

    # ── File i biznesit — smart mapping ──
    bpath = ensure_biznes(b)
    mapping, orig_headers = detect_column_mapping(bpath)

    # Nese file eshte i ri (vetem header-i yne) — perdor format te sistemit
    any_mapped = any(v is not None for v in mapping.values())
    if any_mapped:
        append_row_mapped(bpath, full_entry, mapping)
    else:
        row_b = [dt, ora, entry.get("produkti",""), entry.get("sasia",""),
                 entry.get("njesia",""), entry.get("cmimi",""), entry.get("pagesa",""),
                 entry.get("total",""), entry.get("shenime","")]
        append_row(bpath, row_b, HEADERS_BIZNES)

    log(f"Dergese u ruajt: {b}  |  {entry.get('produkti','')}  |  EUR {entry.get('pagesa','')}")
    return bpath

_FIN_KEYWORDS = ["fitim", "shpenzim", "total", "pagese", "pagesa", "vlera",
                 "cmim", "çmim", "sasi", "shuma", "revenue", "profit",
                 "expense", "amount", "price", "cost", "total", "euro", "eur"]

def _is_financial_col(name: str) -> bool:
    n = name.lower().replace("ë","e").replace("ç","c")
    return any(k in n for k in _FIN_KEYWORDS)

def get_stats():
    """Lexon statistikat direkt nga file-at e bizneseve."""
    _, biz_dir, _ = paths()
    total_rows   = 0
    total_biz    = 0
    fin_cols     = {}   # { col_name: total_value }

    for f in biz_dir.glob("*.xlsx"):
        if SYS_FOLDER in str(f):
            continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            fin_idx = {i: h for i, h in enumerate(headers) if h and _is_financial_col(h)}
            rows = 0
            for r in ws.iter_rows(min_row=2, values_only=True):
                if any(r):
                    rows += 1
                    for i, col_name in fin_idx.items():
                        try:
                            val = float(r[i]) if r[i] is not None else 0
                            fin_cols[col_name] = fin_cols.get(col_name, 0) + val
                        except (TypeError, ValueError):
                            pass
            if rows > 0:
                total_rows += rows
                total_biz  += 1
            wb.close()
        except:
            pass
    return {"dergesa": total_rows, "biznese": total_biz, "fin_cols": fin_cols}

# ── COLUMN PREVIEW WINDOW ────────────────────────────────────────────────────
def show_column_preview(folder_path):
    """
    Shfaq shkurtimisht file-at e lexuara — pa pyetje per timestamp.
    """
    p = Path(folder_path)
    # Prioritet: skanon direkt folder-in e zgjedhur, kapërcen SYS_FOLDER
    files = [f for f in p.glob("*.xlsx") if SYS_FOLDER not in str(f)]
    # Nëse nuk ka direkt, provo nënfolderin Bizneset
    if not files:
        biz_dir = p / "Bizneset"
        files = list(biz_dir.glob("*.xlsx")) if biz_dir.exists() else []
        files = [f for f in files if SYS_FOLDER not in str(f)]
    files = files[:20]

    if not files:
        return  # Asnje file ekzistues — nuk nevojitet preview

    win = tk.Tk()
    win.title("NjoPerKrejt - SmartRegister")
    win.geometry("820x580")
    win.configure(bg=BG)
    win.update_idletasks()
    x = (win.winfo_screenwidth()  - 820) // 2
    y = (win.winfo_screenheight() - 580) // 2
    win.geometry(f"820x580+{x}+{y}")

    tk.Label(win, text=f"U lexuan {len(files)} file Excel",
             font=("Segoe UI", 11, "bold"), bg=BG, fg=WHITE).pack(pady=(16,2))
    tk.Label(win, text="Kolonat e formularëve të tu janë lexuar — asgjë nuk do të ndryshohet.",
             font=("Segoe UI", 9), bg=BG, fg=MUTED).pack()

    # Scrollable area
    container = tk.Frame(win, bg=BG)
    container.pack(fill="both", expand=True, padx=16, pady=10)
    canvas = tk.Canvas(container, bg=BG, highlightthickness=0)
    sb = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=sb.set)
    sb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    inner = tk.Frame(canvas, bg=BG)
    wid = canvas.create_window((0,0), window=inner, anchor="nw")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(wid, width=e.width))
    canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

    for f in sorted(files):
        row = tk.Frame(inner, bg=SURFACE, padx=14, pady=8,
                       highlightbackground=BORDER, highlightthickness=1)
        row.pack(fill="x", pady=(0, 4))
        tk.Label(row, text="📄", font=("Segoe UI", 10),
                 bg=SURFACE, fg=MUTED).pack(side="left", padx=(0, 8))
        tk.Label(row, text=f.stem.replace("_", " "),
                 font=("Segoe UI", 9, "bold"), bg=SURFACE, fg=WHITE).pack(side="left")
        tk.Label(row, text="✔ lexuar", font=("Segoe UI", 8),
                 bg=SURFACE, fg=GREEN).pack(side="right")

    # Buton i thjeshtë Vazhdo
    tk.Button(win, text="  VAZHDO  ▶", font=FONT_B,
              bg=ACCENT, fg="#0d0f14", relief="flat", pady=12,
              cursor="hand2", command=win.destroy).pack(fill="x", padx=16, pady=(8,12))

    win.mainloop()


# ── SETUP WIZARD (here e pare) ────────────────────────────────────────────────
def _detect_biz_count(p: Path) -> int:
    """Numeron sa file .xlsx ka brenda (direkt ose ne Bizneset/)."""
    p = Path(p)
    biz_sub = p / "Bizneset"
    files = list(biz_sub.glob("*.xlsx")) if biz_sub.exists() else []
    if not files:
        files = [f for f in p.glob("*.xlsx") if SYS_FOLDER not in str(f)]
    return len(files)


def run_setup_if_needed():
    cfg = load_config()
    if cfg.get("base_dir"):
        return  # tashmë konfiguruar

    # ── Dritarja e Setup-it ──────────────────────────────────────────────────
    root = tk.Tk()
    root.title("NjoPerKrejt - SmartRegister  |  Konfigurimi i Parë")
    root.geometry("560x420")
    root.configure(bg=BG)
    root.resizable(False, False)
    root.update_idletasks()
    x = (root.winfo_screenwidth()  - 560) // 2
    y = (root.winfo_screenheight() - 420) // 2
    root.geometry(f"560x420+{x}+{y}")

    # ── Titulli ──────────────────────────────────────────────────────────────
    tk.Label(root, text="  NjoPerKrejt", font=("Segoe UI", 15, "bold"),
             bg=BG, fg=WHITE).pack(pady=(24, 2))
    tk.Label(root, text="Mirë se erdhe!  Konfiguro folder-in e punës njëherë e mirë.",
             font=("Segoe UI", 9), bg=BG, fg=MUTED).pack()

    # ── Karta me udhëzime ────────────────────────────────────────────────────
    info = tk.Frame(root, bg="#0d2137", padx=16, pady=12,
                    highlightbackground="#1d4ed8", highlightthickness=1)
    info.pack(fill="x", padx=24, pady=(14, 6))

    steps = (
        "1.  Klikoni  «Zgjidh Folder»  dhe gjeni folder-in ku i keni bizneset.",
        "2.  P.sh.  D:\\Klientet   ose   C:\\Punë\\Bizneset",
        "3.  Klikoni  «FILLO PUNËN»  —  gjithçka tjetër bëhet vetë.",
    )
    for s in steps:
        tk.Label(info, text=s, font=("Segoe UI", 9),
                 bg="#0d2137", fg="#93c5fd", anchor="w", justify="left").pack(
                 fill="x", pady=1)

    # ── Inputi i folder-it ───────────────────────────────────────────────────
    frame = tk.Frame(root, bg=CARD, padx=20, pady=14,
                     highlightbackground=BORDER, highlightthickness=1)
    frame.pack(fill="x", padx=24, pady=6)

    tk.Label(frame, text="Folder-i i zgjedhur:", font=FONT_S, bg=CARD, fg=MUTED).pack(anchor="w")

    row = tk.Frame(frame, bg=CARD)
    row.pack(fill="x", pady=(4, 0))

    path_var   = tk.StringVar(value="")
    status_var = tk.StringVar(value="")

    ent = tk.Entry(row, textvariable=path_var, font=FONT_S,
                   bg=SURFACE, fg=TEXT, insertbackground=WHITE,
                   relief="flat", highlightbackground=BORDER, highlightthickness=1)
    ent.pack(side="left", fill="x", expand=True, ipady=7)

    def _update_status(*_):
        p = path_var.get().strip()
        if not p:
            status_var.set("")
            btn_ok.config(state="disabled", bg="#374151")
            return
        pp = Path(p)
        if not pp.exists():
            status_var.set("⚠  Folder-i nuk ekziston — do të krijohet automatikisht.")
            lbl_status.config(fg=YELLOW)
            btn_ok.config(state="normal", bg=GREEN)
            return
        n = _detect_biz_count(pp)
        if n > 0:
            status_var.set(f"✔  U gjet {n} biznes(e) në këtë folder — gati për punë!")
            lbl_status.config(fg=GREEN)
        else:
            status_var.set("ℹ  Folder bosh — bizneset do të krijohen kur të shtosh të parat.")
            lbl_status.config(fg=MUTED)
        btn_ok.config(state="normal", bg=GREEN)

    path_var.trace_add("write", _update_status)

    def browse():
        p = filedialog.askdirectory(title="Zgjidh folder-in ku i ke bizneset")
        if p:
            pp = Path(p)
            # Nëse klienti zgjodhi direkt folder-in "Bizneset", ngjitu një nivel
            if pp.name.lower() == "bizneset":
                pp = pp.parent
            path_var.set(str(pp))

    tk.Button(row, text="📁  Zgjidh Folder", font=FONT_S, bg=ACCENT, fg=WHITE,
              relief="flat", cursor="hand2", padx=10,
              command=browse).pack(side="right", padx=(8, 0), ipady=7)

    lbl_status = tk.Label(frame, textvariable=status_var,
                          font=("Segoe UI", 8), bg=CARD, fg=MUTED, anchor="w")
    lbl_status.pack(fill="x", pady=(6, 0))

    # ── Butoni Vazhdo ────────────────────────────────────────────────────────
    btn_ok = tk.Button(root, text="FILLO PUNËN  ▶", font=("Segoe UI", 11, "bold"),
                       bg="#374151", fg=WHITE, relief="flat", pady=13,
                       cursor="hand2", state="disabled")
    btn_ok.pack(fill="x", padx=24, pady=(10, 0))

    def confirm():
        p = Path(path_var.get().strip())
        try:
            p.mkdir(parents=True, exist_ok=True)
            # Detekto nëse klienti zgjodhi direkt folder-in me xlsx të bizneseve
            direct_xlsx = [f for f in p.glob("*.xlsx") if SYS_FOLDER not in str(f)]
            has_biz_sub = (p / "Bizneset").exists()
            if direct_xlsx or (not has_biz_sub and p.name.lower() not in ("", "njoperkrejt", "njo per krejt")):
                # Ky folder ËSHTË folder i bizneseve — përdore direkt, mos krijo nënfolder
                base_p = p.parent
                base_p.mkdir(parents=True, exist_ok=True)
                set_dirs(base_p, biz_path=p)
            else:
                # Folder i ri bosh — struktura standarde me nënfolder Bizneset
                set_dirs(p)
                (p / "Bizneset").mkdir(exist_ok=True)
            root.destroy()
            show_column_preview(p)
        except Exception as ex:
            messagebox.showerror("Gabim", f"Folder nuk u krijua:\n{ex}")

    btn_ok.config(command=confirm)

    tk.Label(root, text="Ky konfigurim bëhet vetëm njëherë.  Mund ta ndryshosh më vonë te  Cilësimet.",
             font=("Segoe UI", 8), bg=BG, fg="#4b5563").pack(pady=(8, 0))

    root.mainloop()

# ── APP ───────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NjoPerKrejt - SmartRegister")
        self.geometry("860x660")
        self.minsize(760, 580)
        self.configure(bg=BG)
        self._build()
        self.refresh_stats()

    def _build(self):
        # Header — lartësi dinamike 2 rreshta
        hdr = tk.Frame(self, bg=SURFACE)
        hdr.pack(fill="x")

        # ── E majta: 2 rreshta teksti ──────────────────────────────────────
        left = tk.Frame(hdr, bg=SURFACE)
        left.pack(side="left", padx=16, pady=10)
        tk.Label(left, text="NjoPerKrejt",
                 font=("Segoe UI", 13, "bold"), bg=SURFACE, fg=WHITE).pack(anchor="w")
        tk.Label(left, text="SmartRegister Excel",
                 font=("Segoe UI", 9), bg=SURFACE, fg="#66FFCC").pack(anchor="w")

        # Logo djathtas (SVG me cairosvg)
        try:
            import base64 as _b64, io as _io
            from PIL import Image as _Img, ImageTk as _ITk
            _LOGO_PNG_B64 = "iVBORw0KGgoAAAANSUhEUgAAAN8AAAAWCAYAAABAHklQAAABCGlDQ1BJQ0MgUHJvZmlsZQAAeJxjYGA8wQAELAYMDLl5JUVB7k4KEZFRCuwPGBiBEAwSk4sLGHADoKpv1yBqL+viUYcLcKakFicD6Q9ArFIEtBxopAiQLZIOYWuA2EkQtg2IXV5SUAJkB4DYRSFBzkB2CpCtkY7ETkJiJxcUgdT3ANk2uTmlyQh3M/Ck5oUGA2kOIJZhKGYIYnBncAL5H6IkfxEDg8VXBgbmCQixpJkMDNtbGRgkbiHEVBYwMPC3MDBsO48QQ4RJQWJRIliIBYiZ0tIYGD4tZ2DgjWRgEL7AwMAVDQsIHG5TALvNnSEfCNMZchhSgSKeDHkMyQx6QJYRgwGDIYMZAKbWPz9HbOBQAAAhjklEQVR4nO18eXhV1dX3b+197pDh5pLhZg6ZGAMoQ5hkSkSBwosTJKDAp+IAUi3tZ8Wx3hv1k1ZtleqHWtsqFatNLG2l4gBpgiAgJgJCIJEhTBKSQObh3nvO3vv9454bQoCAffv1a5+nv+e5T3LOPXvvtdc6a6+1117rEgCMHj06+vnnn386LS1tltVqTZZSQtO0A42NjW8NHjz4FSLySimJiBR6QCnFiEgWFRWNmDt37t0ADADapk2btl1//fXvBL/v0YYTkQCAxx9/POXaa68dp+v6xH79+mnp6elgjAEAamtrcfDgQXDOv963b9/We++9t6Jn+0vB7Xazp556SiqlAMCxYcOGkZGRkZMjIiLiMjMzYbPZ6CLNJAB28ODBYwMGDHgOAMrKyu4bNWrU0OB3vY15sb4ArCWi7QCwc+fOjNGjRz8IQADgW7du3Tlp0qQ1SikGQBGRqqqqun/AgAGDAcAwDF9DQ8OTcXFxbUqpi/L/Ygg+axjGQ5zzNAA4ceKEsXjx4ic3bdrUfKm+grLatm3b7PHjx8+AKcu1a9f+ZtGiRV8BQGFh4c15eXmzAbQC4J2dnWrVqlU/f/TRR49eTNYX67+kpCRt4sSJD2qaRiYvHOvWrVs/Z86cP3XvI6+wkBfl5wsA2szfPX173LBBU/zSGCY5C8juohKU0LiG1kNttO/x3YpaZUjy/06zp0/r2+j1ClIWpb6bGP8+BJlLulChjvDyhu37Pv/THY/9HoAvOC9t2bJl4Q8++OD6jIyM8UHyzb9D4uLint+5c+fsMWPGTAfg79EvAKC0tJQBkJzz/gCWBe8LIVwA3kFgpl0CMZkrVq1alTV9+vSHYmNjb4qMjOxzsQnExcUhLi4uQMyQIb7Zs2dvrqqqeoGINjLGIOXF5ex2u1lBQYEEYN+/f//SmJiY7zscjn52u/0KWYcKAM8BgJQyD0DulTa8CL4GsB0AbDZbErrxiIhiAazBOR4pIsoDMBkAOOeIjY1dCaDtO45JZl//C8BQALDb7ZBSPgeg2ePxEHrI0QQDIIlocnc6DcPYBuCrsrIyy/bt27+or69/0uVyDQeAkJAQLFiwIPvtt9+eAkCqgGZf0Lfb7Wacczlv3ryUwYMHb9A0bXDwu9ra2l2nTp16IrgIAecU7/oXH5ycPvu6l8LS4kcY3AoDQZ3rPsS56SgoMFhhtTRCbzoENDPEZ12FyCGZfdvQDt71fK+s6+X6yhCcCIFAwHBLZspdS6eMWlb526InivLzP8krLOTaHXfc8YSpeDoAjnPLggAgR48ePbm4uPhRInL3ZnGOHDkSbbbxArB7vd4LKC4sLOREJMrLy2ekpqb+MTo6OtT8Skfg5bsYVxQAioiIsEVEREyLjY2ddvTo0SfT0tKeLSkpodzcXNGdO263mz399NPygQceSH7kkUfeSUxMnGx+Jc1xeuO8AMA1TWsK3uCcNyOw8AiTP0EQzl9Ce/LFAKAB8HXdMAzDfM4PwMo5b+8xT9TV1YX3799fAIBSqvVKrd3FwBhrDdLl9/u9jLEr6ouIOtFNlia9OHLkCD3wwAOn/va3v01//fXXt7lcrjQppZ6SkjLuvffe+z0RzVVKaebcz01MKQJABQUFlpUrV/4hLi5usJTSyxizHj9+/Oj8+fNnbN++ve7MmTOsoKBAuZViBURi3tvP58TMvGYDi4oJaUGrANoBCerNbimpwJgPRrsPgALIoI5OH1pkqyKjEzpnvb8B/0CQUlBEkIyUnQhaWlr2oGULPyZnxI1F+fkfaBkZGbNwzkVip06dOqzremtqaupwKSUxxuSwYcOWLVy48DkAHZdyWYiIEHg5OQAedB2DKCws5Pn5+WLdunXj+/fv/6HD4WBSSp0xpgGwBJ9TSoGIuv52hxBC55xTamrqU59//rl9woQJj3dfEIJCfu2118IefPDBDxITE0cA8EspmTmOrTdmCSHAOUdzc3Nk8F5LS4sTgNUwDGia1kVjt3kH7/Hu14ZhWDVNQ3Nzs73bs+fxyFzpe/KR4ZyS857ff0d09UVEV9yXlDLYjgPgJt1wuVzSlGPdrFmz8m644YYdLpfLKoTQhw0bNqesrOwpInpSKaURUVABCQAjIrFv377309PTxwsh/Jxze2Nj49lt27ZN2759e13w/YDbzTyA+uNds+PCJgz+s4qKCvEZTQbjTOOKoECQRAAITOoQIEAxkOlckbmE6+Y/XOlgTMLCJAliUEwDQUBBQEkNUArUzbIpEBgUKEA1SDFIEjAkAAVwAjTikCSgAlbtHN+IwBQBkDCkhNXQIJkEMQmdGLxGkxEWH6/1y7v+jbYjJ3ZqHR0dA6KjowkAtba21k2dOnVsZWXl2VOnTq1PSEiYBcDvcrli7rnnnoVE9HpJSckFK9sVgPLy8uTtt99unzBhwm9MxTMYYxav14umpqY/njhx4m+HDh2qqa6uVhaLRem6TjExMRgwYIAjNjZ2YkJCwvzIyMgIIYTknBsjR458eN26dX9hjO3sEpwp5M8++6wgNTV1BAC/EELjnLP6+vqvT506ta6lpWVPeXk5vF6v4pxDiHMGizGmiIjq6+ubgvfWrVv32JdffhmnlFJSSrLb7bylpUVMnDhx6bXXXjsDAWtq2bFjh2fjxo27w8PDmd/vl0oppWka+Xy+3d3c4H975Ofni5KSEi03N3dXcXHxzePHj/8wJCREATCGDx/+ky1btuwjokLzGaOsrEwjIv3IkSPPpKen3ySl1DnnvL293ff+++8vuPfeew8HnwUAtyeHEZFx+2dr7o9MH+hslY260piFAj4cAEBBBrYc3AYNDFZIyG5OCIMGw9EOSYDBLEHXD4o4FPmhhAQnCzTOQF3OnoICmYNI+KFDCh2KabAqHaE8FAQFBR86lYAGDkB2qW139RWQCOE2CA5YwOFTOggA50zrkM0irG9abNacGY9oSikrTMtnsVhOV1ZWnmWMobi4+KX58+f/l6ZpDABSU1N/BOD1nJycXgMdF0NJSQknIqOsrGx2bGzsYAA6Y4y3trZ27Nix49Zp06Z9cJku3v7Zz372izlz5mzIzMzMkFIadrvdOmTIkIeUUnl5eXlQShFjTCxbtix+6NChdwbnxDlnO3fu/NXYsWOX4ULX8LJ4+eWXd1xiTrkAZpjjYPz48RuvueaabRd7trCw8H9qwf6lkJuba5jWbcPGjRtXXHfddc8B8HPORVZW1rt//etfD+fm5paXlJTYs7OzvevXr/9Jenr64wjInRmGwffs2TP/3nvv/aSsrMySnZ2tm12TBzmiACDW13WzF34FFfAyCQwSEoJ0MGgI5+Hw1bfh7KEGyAY/IAiKCJAK3GKF91ArOAhMClOpAEBBSYKdO6AMhaY9ddBP+yB0AUUAKQKBgzmA8MwIOPtGoRUtMMiKhs3HoNcSrH05IsfFBZS/m9kTRLBLAYMUONlwdvNJ+GoVtATAOSkBML0lpoh5oSukx+Vp3ZmqlNJM140RUUlOTs725OTk8VJKvW/fvgO3b98+kzG24Uqijd2Rk5MjAcDlci1GYJFQAFhFRcVj06ZN+0ApZS0tLZX19fWqoqKia0pDhgxRLpeLcnJyGBFVdXR0zFmxYsXnoaGhdgAqISFh+qpVq1xEVF9YWGhVSvlvuumm/4qMjIyCua86dOjQl2PHjl2ilKLy8nLLqFGjZFFREbqP0328iooK2r9/vyoqKhJAYA8JgAW/W7BgAX/nnXcEEZ3nwhqGEeF2u7W8vDxWVFQkg897PB5RVFR0paz6t4G5mFqys7Of3759+8Bx48bdBcAXFRVlHTNmzAcvvPDCuNzc3BOffvpp/uTJk59CwFtSAHh5efljEyZMeL+H4gFuNxGRjB8/OJUY9TdgkCLFSAUcQ0kKmrQgRIZi74tf4tib1ZCnGLx6e7eQHoEgwbhCSLgDbdwbUCwAUiqEahpavm7Arh99jravO+D1Ggj4lAG7BiJYrBbAKZGxaACynhgMZrdh3/qTqP75ftgzHZi5eQ5UkgRJ6lJATSlIJaAoHMaBFmyZ9xm8tTqGPHkVEiZloF22BaL4jIigg6LCEs9TPpjhblMBRX19/TPJyckfBjfqsbGxP1FKfYTvFv4hIpI33HCDw263ZyNArrW+vr7xkUceCYbZ9dzc3F77NFfa3YsWLdqYmZl5IwB/eHi4o1+/ftkAPsrLyyMA6Nev30iTPgkA+/btW20qnnaeoK8QprvYJVqPx4OCggKjJ72apomCggLD4/HwoUOHim7tUVhY+F2H/bdAdnZ20AIuOXr0qCs1NfUGIYTP5XIlzp8/f01sbOxzo0aNes9ms8EwDKVpmnXLli1rJk+evNJsd1F5hA7MsILIcsFeTEhoPBS7HtuOAz+vgt1uAUL8cLgcAKlAZE4BghOEANCqYBUMQeVlnNDRovDF3aXo2O0HhQGhMRZwqwZIQHADmrBAb/JBNVtQ8dO9cCQ7kb40A/3vHISzf6hH6+k2VH9wFP3v64dO6QXj59xdQxHsZMHX71WBNRIiBkVg8N2D0IF2sPNicwqa3aZ6Kl8QwgysbDpx4sSR5OTkDABGSkrKmG3bto1gjH3VbZ91ORAAzJgxY0BkZGQkzChgR0fHjs2bNzchsBZc8fnVnj17igHciEBIHC6XazSAj2AqiBBigjmmta2tzbt79+6tN998s3K73d/Z5fwPLgvl8XikUkqNHz/+znXr1pUmJCQMk1KKpKSk3FtuuSU3LCwMAAxN0ywnT54snDx58mJT8S4pDzKYUiywywvu9ZjQYdVC0Fh+Cod+UwVHuAXR46KQ8eRYhPVVkMoaUEAJWCxA524fSu/eBAYtoJGQ0LgVZ7ceQfuBdvBIO/o/OBjJ8zICURqigNtJQMeJZuxdXgZjbydOFB9F0tIM9Mnqgz4TouErbMOJPx5C+l3pIAsFzxMASdC4BlHfgpo/H4eQhJTr+8KaEoY2wwut28bDfNkvHrU1LR0H4D9z5sxPAUAIIS0WC4uPj39cKYW8vLwrFRABQHJycrTFYuEw912dnZ3VuDBcf0mUlpaCiFRZWVlT9/tZWVnn7XkdDkcwushaWlpaX3vttRoA8Hg8f3fI/j+4NEzPgHbs2NGwefPmWXV1dTWMMSalFGFhYUoEIlrakSNH9qakpCxSgVDxecdDVwIJCQMcNR/VQGvQIF0WDP31NYgZHwFrkh3WZA22JA22FA1avAaeAmiCQCSgoLp0pK3GB9nJEZYchkEPZsGeymDra4M9xQpbXwZKISRdk4HY6cnQfRKqxYCUgCSBxIVpQLgFbeVNOPtZPaxkhZISBIKhBDjZceIvJ9B50A8tmiFzYRp0+GC9MKgNoPcXXyil6He/+93a+vr6Ws65BYCIiYm5aeXKlVcDkN8lkDBp0qTzIqTHjx/X8B0EUF9frwAgISHhiHmLA8CePXtGmdcSAKKjo7v6bGlpodraWk5EFxxb/Af/OBCR3Ldvn/XWW2898cUXXxQAIMaYklIqzjlra2s7+uSTT05UShkej+eKM3XOH8QCBh2tBzshZAciR0QhPCkMfn87fJJDSAlDKhiGAlMW6IYBkAAphnPH8gSLLiEFgFABX20n9LOAfsYP/xkd+hkBOsNQt78G9dvqYCUrLC4NFkbQhULitfEIvdoBo0Xi6LuHzBgqQULCwhiYT6D6D4dBPsA5ORbO0THQBSC1c6kB3XEptxPm3k978cUXO+fNm/drl8v1OADD4XDY5syZcw8R3a+UotLS0u/Mx38kHA5Hi/nvuZDWf/BPRWFhIR86dKj/hz/84ZCJEyeugOmMMcaCUfSYuXPn3khEb5eUlGh/17ELGVDQ4G/1goPDGqVBKQ2KKTCS4DAPC0hBkQFJ5h3TnVRgIBgwoMFikdCPSJRM+giSNAQz4gIvDkE0eGG06DAcAhm3DgicCgrAHqoh5aa+qNzZhLqSGrQfakVIZhj8uhfcakdt6Sk0lTeAhUv0nz8YiiSEooDre5HX8nIunyAiVFRUrGpra2tE4DBcxsbG3v7444+nIhCUuSK30Twf7EJSUtLFl4NLwOVyEQCcOXMmI0gbAAwaNKjSvCYAqK+v7xonIiJCxcXFCaXUeQfj/8roaRVaW1v/pU222+1meXl5asqUKeHLly8vjIyMzJCBvD8OQJNSKpvNFj516tRfv/vuu9eYxxTf+ehFgYGBYCEFHQyGQZCmcgXUDAhELAkCElbBARAMZoCRAA8eiCuCJABcQEUAPFyBhSHwCSewcCBsUChibk3BlMKZiJuVAr/0g3OCHzqSb+mLkMRQiJNeHHv/CIg0MMUAGDjy9jGgmeAc2gex013wKh9sRMAltreXtHxAl/XjRFQ/bNiwv4wePfoOAF6n0xk+d+7cxUTkrq6uvpzyKQA4e/bsWV3XhbnvQ2hoaDq6RSUvh5ycHCilqKKiwtn9/ldffRXMjiEAEELUA+gHQDqdTsfSpUsTCgoKDvWSz/gvBV3Xu7J9GGPK4XD8wwJFUsorUmQz2t2FXtLSyOPxgIjUrl27PkxLS8syDMOvaZq1rq6u8ttvv902YsSIxUII3eFwWHNycv5cWFh4FWPsdG9J2MT4BfmhTAKSSdjiQ6EYoB9qM1N3OKRhAGbWixQSGrODOtohfQSmcSCEQ0BBCxy/Q/p1WNOsuHb9LPg0HZAMigAGBU0qSI2AUIKEQKfsgNW0osJQCEt3IHJaNNp/24Zv//It+t+XBTg42vZ14kzJaSgrkDwvFRTGoQwElFyxi1qZy1qtoqIiKKWoqqrqp+3t7ToC1k9lZmbe+cADD0SkpaVdLnyvAGD9+vXfNDY2NsJU+LCwsHFTpkzpA4B6CvtSICJls9m+15325ubmnd2vfT7fLnNMf1hYmH306NETlVI0e/bsf/WDbgKAxMTE4+a10HXd9uqrr7qUUmQuHt8VXS8wY0wlJCRcdqFzu93Mbreftz/Xdf1iizSZ2Sty7969q4cPHz7ZMAxd0zSto6Ojtbi4+I6RI0fedejQoQ2cc4sQwhcfH+8aM2bMp+np6U4A6lJyF8rHScjz5MUkgaAQMyoG3GZDc0UzDr/3DcJZJGwWDTbNDptmQ4jNjlAKxaE/fQPmJbBwjrDk8ICVhIIMtYLZrfB/q9DR2AxbmAUWB8EaTtDCGVQEgwrlIAAWMNhYGPwkYBBMC8eQfFsGrA4L2ipaUbfxFGJYH1SvPQhR70VYagj6zk2HggBnHJICCWs9hSKASx41dCE/P18opfiiRYuqTpw4URgWFrYAgNfhcKTk5+ffSUSrgvw5b4BzjFXmKtf68ssvlwGYDkCPjY2NfPbZZ28nolXm4T4AqNLS0i6B5OTkqNLS0uAhu/+ZZ57Jdrlc0xGwlpaWlpbWysrKMgAoLy+XAFBTU/N5ZmbmsiA9w4cPX0ZEbzHG9LKyMsuoUaMkAHQfp+d49fX16gqPUf6RYABkRETESTOv1bBYLLbExMRhRFRVXV1t83g8+uVoLy0t7VKwhoYGHhUVBQCyT58+2pQpU+xr166liooKzePxiGA/OTk5QSXlBQUFxuLFi6O60YSQkJAzwLmgFwCUlZVp2dnZ+pYtWx4ZOnToUiGErmkafD4f27Bhw5LbbrvtC6WU5Xvf+96tb7311ua4uLjhUkpfamrqsE2bNv2WiPKD58no4ZHI48daiPNWgBzBe0IDlBKIujENYavK4T1k4OtHd6Pp01qEDHQCSpopZMDZPadxdnMTmBJwZkUiMjMGbaINxDVEXRUJi11Bb+7AjoXbEXN9MjSb6lYD1PUPDKUj7db+sKdaISVBagq69CFufBIihvdBw9Z6HH7nMOzR4ahZfwwKgGt2ImyJkegQLeCMA+q8op4Ak8Hg6/TRZZUPOGf93nzzzddvu+22BTabjQNQQ4YMuRvAL4lIWa3Whi6qAVit1s5g+2DZ0enTp3/bt2/fGeYz8qqrrnr2448/Pk1Ef7gcDW+++eagqVOnvud0OpmZYG2tqan5ZPny5fXmHsIgIqxevfrTgQMHNrhcrj4AjKSkpNEVFRVvjB8//uHs7OyGywzz/x3ffPPN0djYWMCsZbnmmmseWrhw4Wfp6el1V9qHx+PhAHDmzJm2qKgoJYQw7Ha79eqrr36YiJbiXHlYT4jly5cPdDqdeQiUB1m9Xq8RFxf3DQDk5eUFVkjzgHzt2rX3jB07diUAg3MOAJbdu3f/KC8v792ysjJLUVGR/Pjjj1vWrFkzf8mSJZudTmccAF96evotW7ZseYWI7jsvy6WgwKxIotOq03tYg/UqH3yKAM5IQEgNodEMI385Bdvv2A7tRCeq3/oG1GP3xIiBWQX0WCuyCoZDt+hQgmAoA31GOJF4zwAcerES6ot6NH7eAHQdRgQ/gTwZHyTixycgPM0FXSgYXIDrBGYFEhb2Q21JHVq3nsG20mJYdA3CyTHwjn6QaL/A2nVBSiW5Ddaz7ceuSPlM68cWL168ZcaMGZsSEhKmSimNyMjIoeXl5UtGjRr1Wv/+/QeZjysAauzYscFACHJzc4VSinJyctb//ve/P5CYmDhYCGGEh4eHTp069b3Dhw8v83q9H2qadrCqqkoFk5jT0tJARAlJSUnDQ0JC5oeFhQUTq1lbW5sqLy9/nohQVFSE/Pz84P70zMMPP/yCy+V6FoCQUsqsrKy79+zZM6u5ubkwIiJi25EjR/xNTU2wWCznrbjBZOhjx441ff/7398crK74J0ECQGVl5Z9HjRr1dEhICBdCKJfLNeaFF174asWKFe/Y7fadlZWVQghxwT4sSHtjY2MlgEMA0NTU9BmAXLNSQmVnZy/Zv39/aktLS1FDQ0Ojz+frSmJPTEwkpdTIlJSUJU6n02UmvqO5ufn41KlTvw1Ws5g8NtasWTNx+vTpv7JYLFJKqRhjlvLy8tfGjRv3UvfsFTNpuiojI2PBzJkzPw4NDWUA/BMnTlz6ySefHMzOzv5FtyoI5UGpBsDQD9e+F5ExcLifAqVuChoYI0jDj6gpicgtnorDa6rRtPcs0GJAmhmcCoC0S0Smu5B2Xyb6ZDnhlx3QmA2AhFf6MXzlKERnxeLEJ0ehN3ihpMnKbsnbBAa/6ASPtEAHg2I6uFJQXINPdiLz1lS076rDib8cBddDYcR3YtRDIxAyJBpe0YmeVT0wuzY4iXBo2pnK4+9ekfJ1a4vq6upVCQkJ1wVlnpWV9WpNTc1yp9PZz7zHAVBtbe3Z7u9GUVER37x5s3fz5s133njjjSWhoaEhUkpd0zSWkZExGWYB6YABAy5JQLCkCIBWVVX12IIFC3Z2zzMlImlev3D8+PFpKSkpOYwxv5SSpaWlJQBYDmB5enp6rxPdu3fvfgBDLlW9//8C3Wg/MGnSpN8OHDjwHs65X0rJ4+LikuLi4lYAQP/+/Xvt5/Tp088Q0U+UUmzt2rW/Hjhw4Aqn02mXgboqGjx48AwEEsIvCfNZCUA7fvz4GwhkJfHCwkIQkXjllVdSr7322r/GxMQoIYTBObfu2bNnU3Z29n3BYulgX2ZlgyU7O7t4x44dS8eOHftrmPWRkydPfm79+vWVRNSVL1xAgYU6ftrVb96QmbDEltE/3SsadcW4hUHC4BzK6IAtw46rC0ZCQYcU51KnoQKbKQ4OAT98wgDjFkCZJUeMYEgdKXckoe8d6VBCh4LsURx0ri+DKUjoAKdAsREL7OH0EIGrXh6P/iuGwdfmQ2h8KGyRIfAqLzRzr8dkwA0WBHAJKCn0cC3K0nmo8vDmX/7uBYaAzy0QOFa4dMoPkVBK0e23377x6NGjXzPGLFJKw2q1qvj4+EEhISGalFIAsLS3txsbN27cCAAej0cCAetZWFjIb7vtti927949u6ampoMxZkFAWYMFpr19wDm36LquHThw4Mns7OyVJSUlWo+ImULAXTI++OCD2ZWVlZ8BsJq1fOoKxuhEIGe0sbeX04TszjtcQSRV07RgdkeQ3z0DIFIpxZYsWbLiwIEDO03auTmW7zK0dwDwOxyODgA4evSoddGiRSe3bt36iFnxEQxg+HvpywfAYIFl23rgwIHPly1b9nMz4BPcB0fecsstf0pOTo6QUvo459aTJ09+dcstt8xVSnGPxxOUQxeys7N1pZQ2bty43xw4cOA5AFYAht1uV9nZ2X9cvXr1QPP9YgCUBx6q3fh13dE/fHqr79TJhhAebWGkFFMQkqRQGhOGUKJDbxOdUhd+bgiv+fFphvDDEB2GV/iEEsQhJCAkKSEJQoEJASW8uiE6RbvwcV34uRA+blz40QwhmBAKSiiSQhIJSVJIgpBSCa/oFFqKVYQPdgpEMtEpOgUpEoKkUFBCQAmDIBggJCNl1WIsoq6mvnbL7ptPfrqjQQsNDQ0WTUIIcV4Y/yKgw4cP+06ePDk7Kirq44iIiOBPAQQPVXlHR4f88ssvH/rBD35Q3TP/M6iAEyZMKF62bNmYhQsX/njAgAEz+vTpE8/NTcOl0Nra2tzQ0PD5zp07X8rPz99oukEX1BUGE8Pvv//+NgDTN2zYsHTIkCF3RUZGDnI4HNbLzA8AoOt6n8s9I6UMRYBvoQBgGIal9xaA3+/XzDYhZpuwi9COzZs3N2VlZU3atWvXQ0lJSUv69OmTYrFYei0ERuCFRmdnZygApKWlGaY1+WVxcfHpYcOGPR4REXGVzWa7LA8aAvjN/PnzC/bs2WN4PB7yeDyw2+2R8+bN+2tCQsIIAGCM2Wtra08+88wz86urq5sBXLJu0VQuTkSPVFVVDRwwYMCNABAfH6/ddNNN2wzDmACgyu12swIqkG63mxU8VvDFbUSj5U1TX7Enx3yPh0dwYVbgdZX6Xgq9+XQM//OfcenWvstb7UGPYucqBY2WFtV6fN/GfZ9s+EH5j39VmVeYx7Vt27atTk5OVgCoo6OjBsAl9zlEJN1uN5s0adJxt9t9XX5+/qNxcXFTmpubByUkJBw+ffr0wQMHDrw6a9asj3q6H0EEFTA/P79i9erVd+bl5bnuu+++kYmJif1qamoGR0REwDAMYowpXdeJMfZtaGjogU8++eSrhx566JhJX68lTUEFZIx5Z86c+RKA//vGG29cNXPmzAF79+4dHR0dbbvY778QkeScs+rq6mO9sD24N/vU6XR2mm6XppSq7iaL81BRUaEAoKqq6lshxGoiEpxzXllZubN7nz1o948YMeL/jBs37tWnn356pNPpHOH3+1NtNtulzuuk1WplZ8+eDdYUStOVJSIqBLCuqKjo+uTk5JFElMg5Vz37OXPmDAB8UVxcXPKLX/zihEkPud1uEJF8/vnn+7W2tlZ++eWXZUopzhijkydPvvb6668fDP5ESC98UyZN6v7771/wxBNPPBseHq7pui7sdntIWFjYUCKqVEpRQUEBCgoKpFu5WQEVHMGjL82c9e7KnKR+mbnO1MRoaePghqB/xg8h/f2QAOMKPh2njhxt9VXXfVi04MdbASCvMI8X5ReJ/waa9juyN2J9ZwAAAABJRU5ErkJggg=="
            _png_bytes = _b64.b64decode(_LOGO_PNG_B64)
            _img = _Img.open(_io.BytesIO(_png_bytes)).convert("RGBA")
            # Resize proporcionalisht — lartësia 38px (2 rreshta teksti)
            _ow, _oh = _img.size
            _nh = 15
            _nw = int(_ow * _nh / _oh)
            _img = _img.resize((_nw, _nh), _Img.LANCZOS)
            self._logo_img = _ITk.PhotoImage(_img)
            tk.Label(hdr, image=self._logo_img, bg=SURFACE).pack(side="right", padx=(0, 16))
        except Exception:
            tk.Label(hdr, text="noctilux.dev", font=("Segoe UI", 8, "bold"),
                     bg=SURFACE, fg="#66FFCC").pack(side="right", padx=(0, 16))

        # ── Qendra: emri biznesi me badge ──────────────────────────────────
        _biz_name = load_config().get("biznes_name", "")
        center = tk.Frame(hdr, bg=SURFACE)
        center.place(relx=0.5, rely=0.5, anchor="center")
        # Badge background
        badge = tk.Frame(center, bg="#1a3a2e", padx=12, pady=4)
        badge.pack()
        self.lbl_biznes = tk.Label(badge, text=_biz_name,
                                   font=("Segoe UI", 11, "bold"), bg="#1a3a2e", fg="#66FFCC")
        self.lbl_biznes.pack()

        # ── Path bar — folder aktual + ikona folder (jo tekst) ──────────────
        pb = tk.Frame(self, bg="#0a0d14", pady=5)
        pb.pack(fill="x")
        tk.Label(pb, text="Folder aktual:", font=("Segoe UI", 8),
                 bg="#0a0d14", fg=MUTED).pack(side="left", padx=(16, 4))
        self.lbl_path = tk.Label(pb, text=str(get_biz_dir()),
                                  font=("Consolas", 8), bg="#0a0d14", fg=ACCENT)
        self.lbl_path.pack(side="left")
        tk.Button(pb, text="📁", font=("Segoe UI", 11), bg="#0a0d14", fg=MUTED,
                  relief="flat", cursor="hand2", bd=0,
                  command=self.open_settings).pack(side="right", padx=(0, 12))

        # Stats — vetëm BIZNESE + DERGESA
        sf = tk.Frame(self, bg=CARD)
        sf.pack(fill="x")
        self.s_biz = self._stat(sf, "BIZNESE", "0", ACCENT)
        self.s_der = self._stat(sf, "DERGESA", "0", GREEN)

        # Tabs
        style = ttk.Style()
        style.theme_use("default")
        style.configure("T.TNotebook",     background=BG,      borderwidth=0)
        style.configure("T.TNotebook.Tab", background=SURFACE, foreground=MUTED,
                        padding=[18, 8],   font=FONT_S,        borderwidth=0)
        style.map("T.TNotebook.Tab",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", WHITE)])

        nb = ttk.Notebook(self, style="T.TNotebook")
        nb.pack(fill="both", expand=True)

        self.t_form = tk.Frame(nb, bg=BG)
        self.t_biz  = tk.Frame(nb, bg=BG)
        self.t_log  = tk.Frame(nb, bg=BG)

        nb.add(self.t_form, text="  Shto Dergese  ")
        nb.add(self.t_biz,  text="  Bizneset  ")
        nb.add(self.t_log,  text="  Historia  ")
        nb.bind("<<NotebookTabChanged>>", self._tab_change)

        self._build_form()
        self._build_biz()
        self._build_log()

    def _stat(self, parent, label, val, color):
        f = tk.Frame(parent, bg=CARD, padx=22, pady=10)
        f.pack(side="left", fill="x", expand=True)
        tk.Label(f, text=label, font=("Segoe UI", 8),
                 bg=CARD, fg=MUTED).pack(anchor="w")
        lbl = tk.Label(f, text=val, font=("Segoe UI", 20, "bold"),
                       bg=CARD, fg=color)
        lbl.pack(anchor="w")
        return lbl

    def _get_biz_list(self):
        _, biz, _ = paths()
        return sorted([f.stem.replace("_", " ") for f in biz.glob("*.xlsx") if SYS_FOLDER not in str(f)])

    def _get_biz_columns(self, biz_name):
        """Lexon kolonat origjinale te file-it xlsx te biznesit."""
        _, biz, _ = paths()
        path = biz / f"{biz_name.strip().replace(' ', '_')}.xlsx"
        if not path.exists():
            return []
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                v = str(cell.value or "").strip()
                if v:
                    headers.append(v)
            wb.close()
            return headers
        except:
            return []

    def _build_form(self):
        self._last_biznesi  = ""
        self._current_cols  = []   # kolonat origjinale te biznesit aktual
        self.ents           = {}   # { col_index_0based: Entry widget }
        self._form_frame    = None # frame i fushatve — ridrejtohet me biz ndryshim

        wrapper = tk.Frame(self.t_form, bg=BG)
        wrapper.pack(fill="both", expand=True)

        self._canvas = tk.Canvas(wrapper, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(wrapper, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        self._inner = tk.Frame(self._canvas, bg=BG)
        self._win_id = self._canvas.create_window((0, 0), window=self._inner, anchor="nw")
        self._inner.bind("<Configure>",
            lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>",
            lambda e: self._canvas.itemconfig(self._win_id, width=e.width))
        self._canvas.bind_all("<MouseWheel>",
            lambda e: self._canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        outer = tk.Frame(self._inner, bg=BG)
        outer.pack(fill="both", expand=True, padx=28, pady=18)
        self._outer = outer

        tk.Label(outer, text="DERGESE E RE", font=("Segoe UI", 8),
                 bg=BG, fg=MUTED).pack(anchor="w", pady=(0, 8))

        # ── Autocomplete Biznesi ─────────────────────────────────────────────
        top_card = tk.Frame(outer, bg=CARD, padx=24, pady=14,
                            highlightbackground=BORDER, highlightthickness=1)
        top_card.pack(fill="x")

        biz_row = tk.Frame(top_card, bg=CARD)
        biz_row.pack(fill="x")
        tk.Label(biz_row, text="Biznesi  *", font=FONT_S, bg=CARD,
                 fg=MUTED, width=28, anchor="w").pack(side="left")

        self._biz_var = tk.StringVar()

        # Entry per autocomplete (jo Combobox)
        self._biz_entry = tk.Entry(biz_row, textvariable=self._biz_var,
                                    font=FONT, bg=SURFACE, fg=TEXT,
                                    insertbackground=WHITE, relief="flat",
                                    highlightbackground=ACCENT, highlightthickness=1)
        self._biz_entry.pack(side="left", fill="x", expand=True, ipady=6)

        tk.Button(biz_row, text="+ I Ri", font=FONT_S,
                  bg=ACCENT, fg=WHITE, relief="flat", cursor="hand2",
                  padx=8, pady=4,
                  command=self._add_new_biz).pack(side="right", padx=(6, 0))

        # Info label
        self._lbl_cols = tk.Label(top_card,
                                   text="Shkruaj emrin e biznesit ose zgjidh nga lista...",
                                   font=("Segoe UI", 8), bg=CARD, fg=MUTED)
        self._lbl_cols.pack(anchor="w", pady=(6, 0))

        # ── Listbox sugjerime — vendoset NEN fushen e biznesit ───────────────
        # Krijohet si Toplevel per ta vendosur saktesisht nen entry
        self._suggest_win  = None  # Toplevel window
        self._suggest_list = None  # Listbox brenda Toplevel

        # ── Zona dinamike e fushatve ──────────────────────────────────────────
        self._fields_area = tk.Frame(outer, bg=BG)
        self._fields_area.pack(fill="x", pady=(10, 0))

        # Status
        self.lbl_status = tk.Label(outer, text="", font=FONT_S, bg=BG, fg=GREEN)
        self.lbl_status.pack(pady=(8, 0), anchor="w")

        # Bindings autocomplete
        self._biz_var.trace_add("write", self._on_biz_type)
        self._biz_entry.bind("<Down>",     self._suggest_down)
        self._biz_entry.bind("<Return>",   self._on_entry_enter)
        self._biz_entry.bind("<Escape>",   lambda e: self._hide_suggestions())
        self._biz_entry.bind("<FocusOut>", lambda e: self.after(200, self._hide_suggestions))

    def _on_biz_type(self, *args):
        """Kur shkruhet ne entry — shfaq sugjerimet."""
        typed = self._biz_var.get().strip().lower()
        all_biz = self._get_biz_list()

        if not typed:
            self._hide_suggestions()
            return

        matches = [b for b in all_biz if typed in b.lower()]

        if not matches:
            self._hide_suggestions()
            return

        # Match i sakte i vetem — gjenero direkt
        if len(matches) == 1 and matches[0].lower() == typed:
            self._hide_suggestions()
            self._load_biz(matches[0])
            return

        self._show_suggestions(matches[:8])

    def _show_suggestions(self, matches):
        """Hap Toplevel nen entry me listen e sugjerimeve."""
        # Krijo Toplevel nese nuk ekziston
        if self._suggest_win is None or not self._suggest_win.winfo_exists():
            self._suggest_win = tk.Toplevel(self)
            self._suggest_win.overrideredirect(True)   # pa border/title
            self._suggest_win.attributes("-topmost", True)

            # Border i jashtëm
            outer_f = tk.Frame(self._suggest_win, bg=ACCENT)
            outer_f.pack(fill="both", expand=True, padx=1, pady=1)

            self._suggest_list = tk.Listbox(
                outer_f,
                font=FONT,
                bg=CARD, fg=TEXT,
                selectbackground=ACCENT,
                selectforeground=WHITE,
                activestyle="dotbox",
                relief="flat", borderwidth=0,
                highlightthickness=0,
                cursor="hand2"
            )
            self._suggest_list.pack(fill="both", expand=True)
            self._suggest_list.bind("<ButtonRelease-1>", self._on_suggest_click)
            self._suggest_list.bind("<Return>",          self._on_suggest_confirm)
            self._suggest_list.bind("<Escape>",          lambda e: self._hide_suggestions())
            self._suggest_list.bind("<Up>",              self._on_suggest_up)
            self._suggest_list.bind("<Down>",            self._on_suggest_nav_down)
            # Klikimi jasht mbyll listen
            self._suggest_win.bind("<FocusOut>",
                lambda e: self.after(150, self._check_focus_out))

        # Mbush listen
        self._suggest_list.delete(0, "end")
        for m in matches:
            self._suggest_list.insert("end", f"  {m}")

        # Pozicionim sakte nen entry
        self._biz_entry.update_idletasks()
        ex = self._biz_entry.winfo_rootx()
        ey = self._biz_entry.winfo_rooty()
        ew = self._biz_entry.winfo_width()
        eh = self._biz_entry.winfo_height()
        row_h = 28
        h = len(matches) * row_h + 4
        self._suggest_win.geometry(f"{ew}x{h}+{ex}+{ey + eh}")
        self._suggest_win.deiconify()

    def _check_focus_out(self):
        """Mbyll listen vetem nese fokusi nuk eshte tek entry ose lista."""
        try:
            fw = self.focus_get()
            if fw not in (self._biz_entry, self._suggest_list):
                self._hide_suggestions()
        except:
            self._hide_suggestions()

    def _hide_suggestions(self):
        if self._suggest_win and self._suggest_win.winfo_exists():
            self._suggest_win.withdraw()

    def _suggest_down(self, event):
        """Shigjeta ↓ nga entry — shko ne listbox pa fshire tekstin."""
        if self._suggest_win and self._suggest_win.winfo_exists()                 and self._suggest_list.size() > 0:
            self._suggest_list.focus_set()
            self._suggest_list.selection_clear(0, "end")
            self._suggest_list.selection_set(0)
            self._suggest_list.activate(0)
        return "break"   # parandalon cursor-in te zhduket

    def _on_suggest_nav_down(self, event):
        """Navigim ↓ brenda listboxes."""
        cur = self._suggest_list.curselection()
        nxt = (cur[0] + 1) if cur else 0
        if nxt < self._suggest_list.size():
            self._suggest_list.selection_clear(0, "end")
            self._suggest_list.selection_set(nxt)
            self._suggest_list.activate(nxt)
        return "break"

    def _on_suggest_up(self, event):
        """Navigim ↑ — nese jemi ne item 0 kthehu te entry."""
        cur = self._suggest_list.curselection()
        if not cur or cur[0] == 0:
            self._biz_entry.focus_set()
        else:
            prv = cur[0] - 1
            self._suggest_list.selection_clear(0, "end")
            self._suggest_list.selection_set(prv)
            self._suggest_list.activate(prv)
        return "break"

    def _get_selected_suggestion(self):
        sel = self._suggest_list.curselection()
        if sel:
            return self._suggest_list.get(sel[0]).strip()
        return None

    def _on_suggest_click(self, event=None):
        name = self._get_selected_suggestion()
        if name:
            self._biz_var.set(name)
            self._hide_suggestions()
            self._load_biz(name)
            self._biz_entry.focus_set()

    def _on_suggest_confirm(self, event=None):
        self._on_suggest_click()
        return "break"

    def _on_entry_enter(self, event=None):
        """Enter ne entry — nese ka match unik ose item i zgjedhur ne liste."""
        # Nese lista eshte hapur dhe ka selektim
        if self._suggest_win and self._suggest_win.winfo_exists():
            name = self._get_selected_suggestion()
            if name:
                self._biz_var.set(name)
                self._hide_suggestions()
                self._load_biz(name)
                return "break"

        typed = self._biz_var.get().strip()
        all_biz = self._get_biz_list()
        matches = [b for b in all_biz if typed.lower() in b.lower()]
        if len(matches) == 1:
            self._biz_var.set(matches[0])
            self._hide_suggestions()
            self._load_biz(matches[0])
        return "break"

    def _load_biz(self, name):
        cols = self._get_biz_columns(name)
        self._regenerate_fields(name, cols)

    def _on_biz_select(self, event=None):
        name = self._biz_var.get().strip()
        if name:
            self._load_biz(name)

    def _refresh_biz_combo(self):
        pass

    def _regenerate_fields(self, biz_name, columns):
        """Gjenero fushat — 1 kolonë, X menjëherë, buton reload."""
        for w in self._fields_area.winfo_children():
            w.destroy()
        self.ents = {}
        self._current_cols = columns
        self._current_biz_name = biz_name

        if not columns:
            tk.Label(self._fields_area,
                     text="Nuk u lexuan kolona nga ky file. Kontrolloje.",
                     font=FONT_S, bg=BG, fg=YELLOW).pack(anchor="w", pady=8)
            self._lbl_cols.config(text="Asnje kolone u lexua.", fg=YELLOW)
            return

        self._lbl_cols.config(
            text=f"{len(columns)} kolona  ·  {biz_name.replace(' ','_')}.xlsx",
            fg=ACCENT)

        # Fusha te fshehura per kete biznes (session)
        hidden_key = f"_hidden_{biz_name}"
        if not hasattr(self, hidden_key):
            setattr(self, hidden_key, set())

        self._draw_fields(biz_name, columns)

    def _draw_fields(self, biz_name, columns):
        """Vizato fushat — thirret nga _regenerate dhe nga reload."""
        # Pastro vetem fushat (jo header)
        for w in self._fields_area.winfo_children():
            w.destroy()
        self.ents = {}

        hidden_key  = f"_hidden_{biz_name}"
        hidden_set  = getattr(self, hidden_key, set())
        today       = datetime.now().strftime("%d/%m/%Y")

        # ── Wrapper kryesor ──────────────────────────────────────────────────
        form = tk.Frame(self._fields_area, bg=CARD,
                        highlightbackground=BORDER, highlightthickness=1)
        form.pack(fill="x", padx=2, pady=2)

        # ── Toolbar: "Reload fushat" ─────────────────────────────────────────
        toolbar = tk.Frame(form, bg=CARD, padx=16, pady=8)
        toolbar.pack(fill="x")

        hidden_count = len([i for i in range(len(columns)) if i in hidden_set])
        if hidden_count > 0:
            reload_text = f"↺  Rivendos të gjitha  ({hidden_count} të fshehura)"
        else:
            reload_text = "↺  Rivendos fushat"

        self._reload_btn = tk.Button(
            toolbar, text=reload_text,
            font=("Segoe UI", 10), bg=CARD, fg=MUTED,
            relief="flat", cursor="hand2",
            activebackground=CARD, activeforeground=WHITE,
            command=lambda: self._reload_fields(biz_name, columns))
        self._reload_btn.pack(side="right")

        # ── Fushat — 1 kolonë ────────────────────────────────────────────────
        fields_wrap = tk.Frame(form, bg=CARD, padx=16, pady=4)
        fields_wrap.pack(fill="x")

        for idx, col_title in enumerate(columns):
            if idx in hidden_set:
                continue

            row = tk.Frame(fields_wrap, bg=CARD, pady=6)
            row.pack(fill="x")

            # Label + X në të njëjtin rresht
            lbl_row = tk.Frame(row, bg=CARD)
            lbl_row.pack(fill="x")

            tk.Label(lbl_row, text=col_title,
                     font=("Segoe UI", 10), bg=CARD, fg=MUTED,
                     anchor="w").pack(side="left")

            def make_hide(i=idx, r=row, bname=biz_name, cols=columns):
                def hide():
                    getattr(self, f"_hidden_{bname}").add(i)
                    if i in self.ents:
                        del self.ents[i]
                    r.destroy()
                    self._rebind_entries()
                    # Update reload button count
                    hc = len(getattr(self, f"_hidden_{bname}", set()))
                    if hasattr(self, '_reload_btn') and self._reload_btn.winfo_exists():
                        self._reload_btn.config(
                            text=f"↺  Rivendos të gjitha  ({hc} të fshehura)",
                            fg=ACCENT)
                return hide

            x_btn = tk.Button(lbl_row, text="✕",
                              font=("Segoe UI", 9), bg=CARD,
                              fg="#2a4a3e", relief="flat", cursor="hand2",
                              activebackground=CARD, activeforeground=RED,
                              bd=0, padx=4,
                              command=make_hide())
            x_btn.pack(side="right")

            # Entry — i madh, i qartë
            e = tk.Entry(row,
                         font=("Segoe UI", 13),
                         bg=INPUT_BG, fg=INPUT_FG,
                         insertbackground=ACCENT,
                         relief="flat",
                         highlightbackground=BORDER,
                         highlightthickness=2)
            e.pack(fill="x", ipady=10)

            # Focus: border bëhet teal
            e.bind("<FocusIn>",  lambda ev, w=e: w.config(highlightbackground=ACCENT))
            e.bind("<FocusOut>", lambda ev, w=e: w.config(highlightbackground=BORDER))

            col_lower = col_title.lower()
            if any(x in col_lower for x in ["data", "date", "dt", "dita"]):
                e.insert(0, today)

            self.ents[idx] = {"entry": e, "col_title": col_title}

        self._rebind_entries()

        # ── Separator ────────────────────────────────────────────────────────
        tk.Frame(form, bg=BORDER, height=1).pack(fill="x", padx=16, pady=(8, 0))

        # ── Butonat ──────────────────────────────────────────────────────────
        btn_row = tk.Frame(form, bg=CARD, padx=16, pady=12)
        btn_row.pack(fill="x")

        self._btn_save = tk.Button(
            btn_row, text="  RUAJ DËRGESEN  ",
            font=("Segoe UI", 13, "bold"),
            bg=ACCENT, fg="#0d0f14",
            relief="flat", pady=13,
            cursor="hand2", command=self.save)
        self._btn_save.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._btn_save.bind("<Return>", lambda e: self.save())

        tk.Button(btn_row, text="Pastro",
                  font=FONT_S, bg=SURFACE, fg=MUTED,
                  relief="flat", pady=13, cursor="hand2",
                  command=self.clear_form).pack(side="right", ipadx=16)

        self._canvas.yview_moveto(0)

    def _reload_fields(self, biz_name, columns):
        """Rivendos të gjitha fushat — fshin hidden_set."""
        hidden_key = f"_hidden_{biz_name}"
        setattr(self, hidden_key, set())
        self._draw_fields(biz_name, columns)
        # Fokus te fusha e parë
        if self.ents:
            list(self.ents.values())[0]["entry"].focus_set()

    def _rebind_entries(self):
        """Rilidh Enter navigation pas fshehjes se fushave."""
        entries_list = [info["entry"] for info in self.ents.values()
                        if info["entry"].winfo_exists()]
        self._enter_count = 0

        def make_enter_handler(i):
            def handler(event):
                self._enter_count = 0
                if i + 1 < len(entries_list):
                    entries_list[i + 1].focus_set()
                else:
                    self._on_last_field_enter()
                return "break"
            return handler

        for i, ew in enumerate(entries_list[:-1]):
            ew.bind("<Return>", make_enter_handler(i))
        if entries_list:
            entries_list[-1].bind("<Return>", lambda e: (self._on_last_field_enter(), "break")[1])

    def _add_new_biz(self):
        win = tk.Toplevel(self)
        win.title("Biznes i Ri")
        win.geometry("400x160")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text="Emri i Biznesit te Ri:", font=FONT_S,
                 bg=BG, fg=MUTED).pack(anchor="w", padx=24, pady=(20, 4))
        e = tk.Entry(win, font=FONT, bg=SURFACE, fg=TEXT,
                     insertbackground=WHITE, relief="flat",
                     highlightbackground=BORDER, highlightthickness=1)
        e.pack(fill="x", padx=24, ipady=7)
        e.focus_set()

        def confirm():
            name = e.get().strip()
            if not name:
                return
            ensure_biznes(name)
            self._biz_var.set(name)
            log(f"Biznes i ri u shtua: {name}")
            win.destroy()
            # Gjenero formularin per biznesin e ri
            cols = self._get_biz_columns(name)
            self._regenerate_fields(name, cols)
            self._biz_entry.focus_set()

        tk.Button(win, text="Shto Biznesin", font=FONT_B,
                  bg=GREEN, fg=WHITE, relief="flat", pady=10,
                  cursor="hand2", command=confirm).pack(fill="x", padx=24, pady=12)
        win.bind("<Return>", lambda ev: confirm())

    def _on_last_field_enter(self):
        """Enter ne fushen e fundit — here e pare highlight butonin, here e dyte ruaj."""
        self._enter_count += 1
        if self._enter_count == 1:
            # Enter i pare — ndrysho ngjyren e butonit si sinjal vizual
            self._btn_save.config(bg="#16a34a")  # me i ndrite
            self._btn_save.focus_set()
        else:
            # Enter i dyte — ruaj
            self._enter_count = 0
            self._btn_save.config(bg=GREEN)
            self.save()

    def _refresh_biz_dropdown(self):
        # Rifresko sugjerimet nese nevojitet
        if self._last_biznesi:
            self._biz_var.set(self._last_biznesi)

    def save(self):
        biznesi = self._biz_var.get().strip()
        if not biznesi:
            messagebox.showerror("Gabim", "Zgjidh ose shto nje biznes!")
            return
        if not self.ents:
            messagebox.showerror("Gabim", "Formulari eshte bosh. Zgjidh biznesin nga dropdown!")
            return

        # Nderto entry dict nga kolonat origjinale
        _, biz, _ = paths()
        biz_path = biz / f"{biznesi.strip().replace(' ', '_')}.xlsx"

        # Gjej mapping per keto kolona
        col_mapping, orig_headers = detect_column_mapping(biz_path)

        # Mblidh vlerat nga fushat — ruaj me col_index
        col_values = {}
        for idx, info in self.ents.items():
            val = info["entry"].get().strip()
            col_values[idx + 1] = val  # 1-based

        now = datetime.now()
        dt  = ""
        # Gjej fushen date nga col_values
        for idx, info in self.ents.items():
            if any(x in info["col_title"].lower() for x in ["data","date","dt","dita"]):
                dt = info["entry"].get().strip()
                break
        if not dt:
            dt = now.strftime("%d/%m/%Y")
        ora = now.strftime("%H:%M:%S")

        # Nderto produktin per log
        produkti = ""
        for idx, info in self.ents.items():
            if any(x in info["col_title"].lower() for x in ["produkt","sherbim","item","mall","artikull"]):
                produkti = info["entry"].get().strip()
                break

        # Shto ne file te biznesit — SAKTESISHT sipas kolonave te tij
        try:
            wb = openpyxl.load_workbook(biz_path)
            ws = wb.active
            nr = ws.max_row + 1
            H_BG_loc = H_BG
            for idx, info in self.ents.items():
                col_num = idx + 1
                val = info["entry"].get().strip()
                # Konverto numra
                try:
                    if val and not any(c.isalpha() for c in val):
                        val = float(val.replace(",", "."))
                except:
                    pass
                cell = ws.cell(row=nr, column=col_num, value=val)
                cell.alignment = Alignment(vertical="center")
                if nr % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=ROW_ALT)
                cell.border = Border(bottom=Side(style="hair", color="CCCCCC"))
            # Shto Data/Ora nëse përdoruesi ka zgjedhur "Po"
            if load_config().get("add_timestamp", False):
                ts_label = "Data/Ora e shtimit"
                ts_col = None
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(row=1, column=c).value or "").strip() == ts_label:
                        ts_col = c
                        break
                if ts_col is None:
                    ts_col = ws.max_column + 1
                    hcell = ws.cell(row=1, column=ts_col, value=ts_label)
                    hcell.font = Font(bold=True, color="FFFFFF", size=11)
                    hcell.fill = PatternFill("solid", fgColor="1e3a5f")
                    hcell.alignment = Alignment(horizontal="center", vertical="center")
                ts_val = f"{dt}  {ora}".strip()
                cell = ws.cell(row=nr, column=ts_col, value=ts_val)
                cell.alignment = Alignment(vertical="center")
                if nr % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=ROW_ALT)
            auto_width(ws)
            wb.save(biz_path)
        except Exception as ex:
            messagebox.showerror("Gabim ne ruajtje", str(ex))
            return

        log(f"Dergese u ruajt: {biznesi}  |  {produkti}  |  {biz_path.name}")
        self._last_biznesi = biznesi
        self.refresh_stats()
        self._refresh_biz_dropdown()
        self.clear_form()
        self.lbl_status.config(
            text=f"U ruajt: {biznesi}  |  {produkti}  |  {biz_path.name}")

    def clear_form(self):
        """Pastro vlerat e fushatve, mbaj biznesin e fundit"""
        today = datetime.now().strftime("%d/%m/%Y")
        for idx, info in self.ents.items():
            info["entry"].delete(0, "end")
            col_lower = info["col_title"].lower()
            if any(x in col_lower for x in ["data", "date", "dt", "dita"]):
                info["entry"].insert(0, today)
        if self._last_biznesi:
            self._biz_var.set(self._last_biznesi)
        self._enter_count = 0
        if hasattr(self, "_btn_save"):
            self._btn_save.config(bg=GREEN)
        # Fokusi ne fushen e pare
        if self.ents:
            self.ents[0]["entry"].focus_set()

    # ── BIZNESET ──────────────────────────────────────────────────────────────
    def _build_biz(self):
        top = tk.Frame(self.t_biz, bg=BG)
        top.pack(fill="x", padx=24, pady=(18, 8))
        tk.Label(top, text="BIZNESET E REGJISTRUARA", font=("Segoe UI", 8),
                 bg=BG, fg=MUTED).pack(side="left")
        tk.Button(top, text="Rifresko", font=FONT_S, bg=SURFACE, fg=MUTED,
                  relief="flat", cursor="hand2",
                  command=self.refresh_biz).pack(side="right")
        tk.Button(top, text="Hap Dosjen", font=FONT_S, bg=SURFACE, fg=ACCENT,
                  relief="flat", cursor="hand2",
                  command=self._open_biz_folder).pack(side="right", padx=6)

        style = ttk.Style()
        style.configure("B.Treeview", background=CARD, foreground=TEXT,
                        fieldbackground=CARD, font=FONT, rowheight=30, borderwidth=0)
        style.configure("B.Treeview.Heading", background=SURFACE,
                        foreground=MUTED, font=FONT_S, relief="flat")
        style.map("B.Treeview", background=[("selected", ACCENT)])

        fr = tk.Frame(self.t_biz, bg=BG)
        fr.pack(fill="both", expand=True, padx=24, pady=(0, 20))

        cols = ("biznesi", "dergesa", "pagesa", "file")
        self.tree = ttk.Treeview(fr, columns=cols, show="headings", style="B.Treeview")
        self.tree.heading("biznesi", text="  Biznesi")
        self.tree.heading("dergesa", text="Dergesa")
        self.tree.heading("pagesa",  text="Totali (EUR)")
        self.tree.heading("file",    text="File Excel")
        self.tree.column("biznesi", width=240, anchor="w")
        self.tree.column("dergesa", width=80,  anchor="center")
        self.tree.column("pagesa",  width=130, anchor="e")
        self.tree.column("file",    width=260, anchor="w")

        sb = ttk.Scrollbar(fr, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self._open_biz_file)
        self.refresh_biz()

    def refresh_biz(self):
        _, biz, _ = paths()
        for r in self.tree.get_children():
            self.tree.delete(r)
        for f in sorted(f for f in biz.glob("*.xlsx") if SYS_FOLDER not in str(f)):
            name = f.stem.replace("_", " ")
            try:
                wb   = openpyxl.load_workbook(f, data_only=True, read_only=True)
                ws   = wb.active
                rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
                cnt  = len(rows)
                # pagesa eshte col index 6 (0-based) ne file biznesit
                tot  = sum(float(r[6] or 0) for r in rows if len(r) > 6 and r[6])
                wb.close()
            except:
                cnt, tot = 0, 0.0
            self.tree.insert("", "end",
                values=(f"  {name}", cnt, f"EUR {tot:,.2f}", f.name))

    def _open_biz_file(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        _, biz, _ = paths()
        fname = self.tree.item(sel[0])["values"][3]
        path  = biz / fname
        if path.exists():
            if sys.platform == "win32":
                os.startfile(str(path))
            else:
                os.system(f'open "{path}"')

    def _open_biz_folder(self):
        _, biz, _ = paths()
        if sys.platform == "win32":
            os.startfile(str(biz))
        else:
            os.system(f'open "{biz}"')

    # ── HISTORIA ──────────────────────────────────────────────────────────────
    def _build_log(self):
        top = tk.Frame(self.t_log, bg=BG)
        top.pack(fill="x", padx=24, pady=(18, 8))
        tk.Label(top, text="HISTORIA E VEPRIMEVE", font=("Segoe UI", 8),
                 bg=BG, fg=MUTED).pack(side="left")
        tk.Button(top, text="Rifresko", font=FONT_S, bg=SURFACE, fg=MUTED,
                  relief="flat", cursor="hand2",
                  command=self.refresh_log).pack(side="right")

        fr = tk.Frame(self.t_log, bg=BG)
        fr.pack(fill="both", expand=True, padx=24, pady=(0, 20))

        self.log_txt = tk.Text(fr, font=("Consolas", 9), bg=CARD, fg=TEXT,
                                relief="flat", state="disabled", wrap="word",
                                highlightbackground=BORDER, highlightthickness=1)
        sb = ttk.Scrollbar(fr, orient="vertical", command=self.log_txt.yview)
        self.log_txt.configure(yscrollcommand=sb.set)
        self.log_txt.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.refresh_log()

    def refresh_log(self):
        _, _, lg = paths()
        self.log_txt.config(state="normal")
        self.log_txt.delete("1.0", "end")
        if lg.exists():
            lines = lg.read_text(encoding="utf-8").strip().split("\n")
            self.log_txt.insert("1.0", "\n".join(reversed(lines)))
        else:
            self.log_txt.insert("1.0", "Ende nuk ka veprime.")
        self.log_txt.config(state="disabled")

    # ── CILESIMET ─────────────────────────────────────────────────────────────
    def open_settings(self):
        win = tk.Toplevel(self)
        win.title("Nderro Folderin")
        win.geometry("520x240")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text="NDERRO FOLDERIN", font=FONT_H,
                 bg=BG, fg=WHITE).pack(pady=(20, 12))

        frame = tk.Frame(win, bg=CARD, padx=20, pady=16,
                         highlightbackground=BORDER, highlightthickness=1)
        frame.pack(fill="x", padx=24)

        tk.Label(frame, text="Folder ku ruhen te dhenat:",
                 font=FONT_S, bg=CARD, fg=MUTED).pack(anchor="w")

        row = tk.Frame(frame, bg=CARD)
        row.pack(fill="x", pady=(6, 0))

        path_var = tk.StringVar(value=str(get_base_dir()))
        ent = tk.Entry(row, textvariable=path_var, font=FONT_S,
                       bg=SURFACE, fg=TEXT, insertbackground=WHITE,
                       relief="flat", highlightbackground=BORDER, highlightthickness=1)
        ent.pack(side="left", fill="x", expand=True, ipady=6)

        def browse():
            p = filedialog.askdirectory(title="Zgjidh Folder")
            if p:
                path_var.set(p)

        tk.Button(row, text="Shfleto", font=FONT_S, bg=ACCENT, fg=WHITE,
                  relief="flat", cursor="hand2", padx=10,
                  command=browse).pack(side="right", padx=(6, 0), ipady=6)

        def apply():
            p = Path(path_var.get().strip())
            try:
                p.mkdir(parents=True, exist_ok=True)
                # Detekto nëse klienti zgjodhi direkt folder-in me xlsx të bizneseve
                direct_xlsx = [f for f in p.glob("*.xlsx") if SYS_FOLDER not in str(f)]
                has_biz_sub = (p / "Bizneset").exists()
                if direct_xlsx or (not has_biz_sub and p.parent != p):
                    # Ky folder ËSHTË folder i bizneseve — përdore direkt
                    base_p = p.parent
                    base_p.mkdir(parents=True, exist_ok=True)
                    set_dirs(base_p, biz_path=p)
                else:
                    set_dirs(p)
                    (p / "Bizneset").mkdir(exist_ok=True)
                self.lbl_path.config(text=str(get_biz_dir()))
                self.refresh_stats()
                self.refresh_biz()
                log(f"Folder u ndryshua ne: {p}")
                win.destroy()
            except Exception as ex:
                messagebox.showerror("Gabim", str(ex))

        btn_row2 = tk.Frame(win, bg=BG)
        btn_row2.pack(fill="x", padx=24, pady=(12,4))
        tk.Button(btn_row2, text="RUAJ", font=FONT_B,
                  bg=GREEN, fg=WHITE, relief="flat", pady=10,
                  cursor="hand2", command=apply).pack(side="left", fill="x", expand=True, padx=(0,6))


        # ── Butoni Reset Setup ───────────────────────────────────────────
        sep = tk.Frame(win, bg=BORDER, height=1)
        sep.pack(fill="x", padx=24, pady=(10, 0))

        def reset_setup():
            # Dritarja e fjalëkalimit
            pw_win = tk.Toplevel(win)
            pw_win.title("Autorizim")
            pw_win.geometry("340x180")
            pw_win.configure(bg=BG)
            pw_win.resizable(False, False)
            pw_win.grab_set()
            pw_win.update_idletasks()
            x = (pw_win.winfo_screenwidth()  - 340) // 2
            y = (pw_win.winfo_screenheight() - 180) // 2
            pw_win.geometry(f"340x180+{x}+{y}")

            tk.Label(pw_win, text="🔒  Kërkohet fjalëkalimi i adminit",
                     font=FONT_B, bg=BG, fg=WHITE).pack(pady=(22, 6))

            pw_var = tk.StringVar()
            err_var = tk.StringVar()

            pw_entry = tk.Entry(pw_win, textvariable=pw_var, show="●",
                                font=("Segoe UI", 11), bg=SURFACE, fg=WHITE,
                                insertbackground=WHITE, relief="flat",
                                highlightbackground=BORDER, highlightthickness=1,
                                justify="center")
            pw_entry.pack(fill="x", padx=30, ipady=8)
            pw_entry.focus_set()

            tk.Label(pw_win, textvariable=err_var,
                     font=("Segoe UI", 8), bg=BG, fg=RED).pack(pady=(4, 0))

            def _do_reset(event=None):
                if pw_var.get() != "admin":
                    err_var.set("✖  Fjalëkalim i gabuar")
                    pw_entry.delete(0, "end")
                    return
                pw_win.destroy()
                try:
                    cfg = load_config()
                    cfg.pop("base_dir", None)
                    cfg.pop("biz_dir", None)
                    cfg.pop("biznes_name", None)
                    cfg.pop("add_timestamp", None)
                    save_config(cfg)
                    win.destroy()
                    self.destroy()
                    run_setup_if_needed()
                    App().mainloop()
                except Exception as ex:
                    messagebox.showerror("Gabim", str(ex))

            pw_entry.bind("<Return>", _do_reset)
            tk.Button(pw_win, text="VAZHDO", font=FONT_B,
                      bg=RED, fg=WHITE, relief="flat", pady=8,
                      cursor="hand2", command=_do_reset).pack(fill="x", padx=30, pady=(8, 0))

        tk.Button(win, text="⚙  Rikonfiguro — Fillo Setup Sërisht",
                  font=FONT_S, bg="#1a1d26", fg="#ef4444",
                  relief="flat", pady=8, cursor="hand2",
                  command=reset_setup).pack(fill="x", padx=24, pady=(6, 14))

        win.geometry("520x320")

    # ── HELPERS ───────────────────────────────────────────────────────────────
    def refresh_stats(self):
        try:
            s = get_stats()
            self.s_biz.config(text=str(s["biznese"]))
            self.s_der.config(text=str(s["dergesa"]))
        except:
            pass

    def _tab_change(self, event):
        tab = event.widget.tab("current", "text").strip()
        if "Bizneset" in tab:
            self.refresh_biz()
        elif "Historia" in tab:
            self.refresh_log()
        elif "Dergese" in tab:
            self._refresh_biz_dropdown()
        self.refresh_stats()


def run_biznes_name_if_needed():
    """Pyet per emrin e biznesit nese nuk eshte vendosur akoma."""
    cfg = load_config()
    if cfg.get("biznes_name", "").strip():
        return  # tasme konfiguruar

    root = tk.Tk()
    root.title("NjoPerKrejt - SmartRegister")
    root.geometry("420x220")
    root.configure(bg=BG)
    root.resizable(False, False)
    root.update_idletasks()
    x = (root.winfo_screenwidth()  - 420) // 2
    y = (root.winfo_screenheight() - 220) // 2
    root.geometry(f"420x220+{x}+{y}")

    tk.Label(root, text="Emri i Biznesit Tuaj",
             font=("Segoe UI", 13, "bold"), bg=BG, fg=WHITE).pack(pady=(28, 4))
    tk.Label(root, text="Ky emer do te shfaqet ne krye te programit.",
             font=("Segoe UI", 9), bg=BG, fg=MUTED).pack()

    frame = tk.Frame(root, bg=CARD, padx=20, pady=14,
                     highlightbackground=BORDER, highlightthickness=1)
    frame.pack(fill="x", padx=24, pady=14)

    name_var = tk.StringVar()
    ent = tk.Entry(frame, textvariable=name_var, font=("Segoe UI", 12),
                   bg=SURFACE, fg=WHITE, insertbackground=WHITE,
                   relief="flat", highlightbackground=BORDER, highlightthickness=1,
                   justify="center")
    ent.pack(fill="x", ipady=9)
    ent.focus_set()

    def confirm(event=None):
        name = name_var.get().strip()
        if not name:
            ent.config(highlightbackground=RED)
            return
        cfg = load_config()
        cfg["biznes_name"] = name
        save_config(cfg)
        root.destroy()

    ent.bind("<Return>", confirm)
    tk.Button(root, text="VAZHDO  ▶", font=("Segoe UI", 10, "bold"),
              bg=GREEN, fg=WHITE, relief="flat", pady=10,
              cursor="hand2", command=confirm).pack(fill="x", padx=24)

    root.mainloop()


if __name__ == "__main__":
    run_setup_if_needed()
    run_biznes_name_if_needed()
    check_for_update()
    App().mainloop()
